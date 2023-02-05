import discord
import datetime
import random
from discord import app_commands
import discord, asyncio
from discord.ext import commands, tasks
from discord import Interaction
from discord.ui import Button, View
from discord.ext import commands
from discord.utils import get
from discord import Intents
from aiohttp import request
import os
from openpyxl import load_workbook, Workbook
import time
from apscheduler.schedulers.background import BackgroundScheduler
from typing import List

token = 'os.environ["BOT_TOKEN"]'

############################게임 확률 명령어###############################

suc = -1
suc1 = -1
per = 0
per1 = 0

def play_game():
    global suc
    per = random.randrange(1, 10001)
    if suc1 == 0:
        if 0 < per <= 5000:
            suc = 0
            print(f'실패')

        elif 5000 < per <=9000:
            suc = 1
            print(f'싱글킬 성공')

        elif 9000 < per <=9700:
            suc = 2
            print(f'더블킬 성공')

        elif 9700 < per <=9900:
            suc = 3
            print(f'클러치 성공')

        elif 9900 < per <=9990:
            suc = 4
            print(f'팀에이스 성공')

        elif 9990 < per <=9999:
            suc = 5
            print(f'에이스 성공')

        elif 9999 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
    elif suc1 == 1:
        if 0 < per <= 4000:
            suc = 0
            print(f'실패')

        elif 4000 < per <=8800:
            suc = 1
            print(f'싱글킬 성공')

        elif 8800 < per <=9500:
            suc = 2
            print(f'더블킬 성공')

        elif 9500 < per <=9800:
            suc = 3
            print(f'클러치 성공')

        elif 9800 < per <=9900:
            suc = 4
            print(f'팀에이스 성공')

        elif 9900 < per <=9990:
            suc = 5
            print(f'에이스 성공')

        elif 9990 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
    elif suc1 == 2:
        if 0 < per <= 3500:
            suc = 0
            print(f'실패')

        elif 3500 < per <=7000:
            suc = 1
            print(f'싱글킬 성공')

        elif 7000 < per <=8500:
            suc = 2
            print(f'더블킬 성공')

        elif 8500 < per <=9300:
            suc = 3
            print(f'클러치 성공')

        elif 9300 < per <=9600:
            suc = 4
            print(f'팀에이스 성공')

        elif 9600 < per <=9850:
            suc = 5
            print(f'에이스 성공')

        elif 9850 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
    elif suc1 == 3:
        if 0 < per <= 3000:
            suc = 0
            print(f'실패')

        elif 3000 < per <=6000:
            suc = 1
            print(f'싱글킬 성공')

        elif 6000 < per <=8000:
            suc = 2
            print(f'더블킬 성공')

        elif 8000 < per <=9000:
            suc = 3
            print(f'클러치 성공')

        elif 9000 < per <=9500:
            suc = 4
            print(f'팀에이스 성공')

        elif 9500 < per <=9800:
            suc = 5
            print(f'에이스 성공')

        elif 9800 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
    elif suc1 == 4:
        if 0 < per <= 2000:
            suc = 0
            print(f'실패')

        elif 2000 < per <=5000:
            suc = 1
            print(f'싱글킬 성공')

        elif 5000 < per <=7000:
            suc = 2
            print(f'더블킬 성공')

        elif 7000 < per <= 8500:
            suc = 3
            print(f'클러치 성공')

        elif 8500 < per <=9300:
            suc = 4
            print(f'팀에이스 성공')

        elif 9300 < per <=9700:
            suc = 5
            print(f'에이스 성공')

        elif 9700 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
    elif suc1 == 5:
        if 0 < per <= 1000:
            suc = 0
            print(f'실패')

        elif 1000 < per <=4000:
            suc = 1
            print(f'싱글킬 성공')

        elif 4000 < per <=6000:
            suc = 2
            print(f'더블킬 성공')

        elif 6000 < per <=8000:
            suc = 3
            print(f'클러치 성공')

        elif 8000 < per <=9000:
            suc = 4
            print(f'팀에이스 성공')

        elif 9000 < per <=9600:
            suc = 5
            print(f'에이스 성공')

        elif 9600 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
    elif suc1 == 6:
        if 0 < per <= 0:
            suc = 0
            print(f'실패')

        elif 0 < per <=2000:
            suc = 1
            print(f'싱글킬 성공')

        elif 2000 < per <=5000:
            suc = 2
            print(f'더블킬 성공')

        elif 5000 < per <=7000:
            suc = 3
            print(f'클러치 성공')

        elif 7000 < per <=8500:
            suc = 4
            print(f'팀에이스 성공')

        elif 8500 < per <=9300:
            suc = 5
            print(f'에이스 성공')

        elif 9300 < per <=10000:
            suc = 6
            print(f'퍼펙트 에이스 성공')
            
def operater_choice():
    global suc1
    per1 = random.randrange(1, 10001)
    if 0 < per1 <= 5000:
        suc1 = 0
        print(f'일반 오퍼')

    elif 5000 < per1 <=7000:
        suc1 = 1
        print(f'보병대 오퍼')

    elif 7000 < per1 <=9000:
        suc1 = 2
        print(f'아이온 오퍼')

    elif 9000 < per1 <=9500:
        suc1 = 3
        print(f'약탈자 오퍼')

    elif 9500 < per1 <=9800:
        suc1 = 4
        print(f'빛의 감시자 오퍼')

    elif 9800 < per1 <=9990:
        suc1 = 5
        print(f'엘더플라임 오퍼')

    elif 9990 < per1 <=10000:
        suc1 = 6
        print(f'아락시스 오퍼')


###############################점수 추가 명령어######################################

def add_score(_name, row, add_amount):
    loadFile()
    print(str(_name), "의 점수를 추가합니다.")
    ws.cell(row, c_score).value += add_amount
    print(_name, "의 점수 : " + str(ws.cell(row, c_score).value))
    print("------------------------------\n")
    saveFile()

def day_count_delete(_name, row):
    loadFile()
    ws.cell(row, c_day_count).value -= 1
    print(_name, "의 횟수 : " + str(ws.cell(row, c_day_count).value))
    saveFile()

#################################메인 명령어################################

class aclient(discord.Client):
    def __init__(self):
        super().__init__(intents = discord.Intents.default())
        self.synced = False

    async def on_ready(self):
        await self.wait_until_ready()
        guild = client.get_guild(1053299673206640711)
        po = guild.member_count
        if not self.synced: 
            await tree.sync() 
            self.synced = True
        print(f'{self.user}이 시작되었습니다')  #  봇이 시작하였을때 터미널에 뜨는 말
        game = discord.Game(f'v0.1.5 beta')          # ~~ 하는중
        await self.change_presence(status=discord.Status.idle, activity=game)


client = aclient()
tree = app_commands.CommandTree(client)
sched = BackgroundScheduler()
wb = load_workbook(r"C:\Users\user\python project\discord.py\main\userDB.xlsx")
ws = wb.active

############################회원가입 명령어#################################

c_name = 1
c_id = 2
c_score = 3
c_day_count = 4
c_rsp = 5
c_vsp = 6

def checkRow():
    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break
def checkName(_name, _id):
    for row in range(2, ws.max_row+1):
        if ws.cell(row,1).value == _name and ws.cell(row,2).value == _id:
            break
            return False
        else:
            return True
            break

    saveFile()

def loadFile():
    wb = load_workbook(r"C:\Users\user\python project\discord.py\main\userDB.xlsx")
    ws = wb.active

def saveFile():
    wb.save(r"C:\Users\user\python project\discord.py\main\userDB.xlsx")
    wb.close()

def checkFirstRow():
    print("checkFirstRow")
    loadFile()

    print("첫번째 빈 곳을 탐색")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUserNum():
    print("checkUserNum")
    loadFile()

    count = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count

def checkUser_name(_name):
    print("checkUser")
    print(str(_name) + "의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        if ws.cell(row, c_name).value == _name:
            print("등록된  이름을 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def checkUser(_name, _id):
    print("checkUser")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def Signup(_name, _id):
    print("signup")

    loadFile()

    _row = checkFirstRow()
    print("첫번째 빈곳: ", _row)
    print("")

    print("데이터 추가 시작")

    ws.cell(row=_row, column=c_name, value=_name)
    print("이름 추가 | ",  ws.cell(_row,c_name).value)
    ws.cell(row=_row, column=c_id, value =hex(_id))
    print("고유번호 추가 | ", ws.cell(_row,c_id).value)
    ws.cell(row=_row, column=c_score, value= 0)
    print("점수 추가 | ", ws.cell(_row,c_score).value)
    ws.cell(row=_row, column=c_day_count, value= 1)
    print("일일 횟수 추가 | ", ws.cell(_row,c_day_count).value)
    ws.cell(row=_row, column=c_rsp, value= 0)
    print("rsp 포인트 추가 | ", ws.cell(_row,c_rsp).value)
    ws.cell(row=_row, column=c_vsp, value= 0)
    print("vsp 포인트 추가 | ", ws.cell(_row,c_vsp).value)

    print("")

    saveFile()

    print("데이터 추가 완료")

@tree.command(name = '회원가입', description='게임 참여를 위해 회원가입을 진행합니다.') 
async def login(interaction: discord.Interaction):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        print("DB에서 ", interaction.user.name, "을 찾았습니다.")
        print("------------------------------\n")
        await interaction.response.send_message(f"이미 회원가입이 되어있습니다.", ephemeral = True) 
    else:
        print("DB에서 ", interaction.user.name, "을 찾을 수 없습니다")
        print("")

        Signup(interaction.user.name, interaction.user.id)

        print("회원가입이 완료되었습니다.")
        print("------------------------------\n")
        await interaction.response.send_message(f"회원가입이 완료되었습니다.", ephemeral = True)

#####################################오퍼레이터 사격 게임#################################

@tree.command(name = '오퍼레이터사격', description='오퍼레이터 사격 게임을 시작합니다.') 
async def Game_start(interaction: discord.Interaction):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    button1 = Button(label="시작하기", style = discord.ButtonStyle.green)
    async def button_callback1(interaction:discord.Interaction):
        if userExistance:
            if int(ws.cell(userRow, c_day_count).value) >= 1:
                await interaction.response.send_message(f'ㅤ', ephemeral = True)
                operater_choice()
                play_game()
                day_count_delete(interaction.user.name, userRow)
                iun = interaction.user.name
                iua = interaction.user.avatar
                if suc1 == 0:
                    embed=discord.Embed(title=f'일반 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070643197392457738/standard-operator_valorant_full_skin_154803.png")
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    msg = await interaction.channel.send(embed=embed)
                    await msg.add_reaction("✅")
                if suc1 == 1:
                    embed=discord.Embed(title=f'보병대 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814><:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070639922404409344/latest.png")
                    msg = await interaction.channel.send(embed=embed)
                    await msg.add_reaction("✅")
                if suc1 == 2:
                    embed=discord.Embed(title=f' 아이온 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070638856673697802/latest.png")
                    msg = await interaction.channel.send(embed=embed)
                    await msg.add_reaction("✅")
                if suc1 == 3:
                    embed=discord.Embed(title=f'약탈자 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070638960293974076/latest.png")
                    msg = await interaction.channel.send(embed=embed)                  
                    await msg.add_reaction("✅")
                if suc1 == 4:
                    embed=discord.Embed(title=f'빛의 감시자 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070638274621739068/latest.png")
                    msg = await interaction.channel.send(embed=embed)
                    await msg.add_reaction("✅")
                if suc1 == 5:
                    embed=discord.Embed(title=f'엘더프라임 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070639690346143774/latest.png")
                    msg = await interaction.channel.send(embed=embed)
                    await msg.add_reaction("✅")
                if suc1 == 6:
                    embed=discord.Embed(title=f'아락시스 오퍼레이터를 획득하셨습니다!',description="✅를 눌러 오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
                    embed.add_field(name="<:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814><:star2:1061560043914399814>",value=f'', inline=False)
                    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
                    embed.set_image(url=f"https://cdn.discordapp.com/attachments/1054445219120164874/1070640081611800576/araxys-operator_valorant_full_skin_609655.png")
                    msg = await interaction.channel.send(embed=embed)
                    await msg.add_reaction("✅")
                @client.event
                async def on_reaction_add(reaction, user):
                    user_name = str(user)
                    print(user_name[-5:-1] + user_name[-1])
                    user_name.strip(user_name[-5:-1] + user_name[-1])
                    username = user_name
                    print(username)
                    if user.bot == 1:
                        return None
                    if interaction.user.name == user:
                        if str(reaction.emoji) == "✅":
                            await msg.delete()
                            if suc == 0:
                                view.clear_items()
                                embed = discord.Embed(title='실패..', description="적이 도망갔습니다.",colour=discord.Colour.red())
                                await interaction.channel.send(embed=embed, view=view)
                            elif suc == 1:
                                view.clear_items()
                                add_score(interaction.user.name, userRow, 1)
                                embed = discord.Embed(title='싱글킬!', description="짧게 피킹 하는 적을 처치했습니다.",colour=discord.Colour.blue())
                                await interaction.channel.send(embed =embed, view=view)
                            elif suc == 2:
                                view.clear_items()
                                add_score(interaction.user.name, userRow, 2)
                                await interaction.channel.send(embed = discord.Embed(title='더블킬!', 
                                description="A롱에서 들어오는 적 두명을 처치했습니다.",colour=discord.Colour.blue()), 
                                view=view)
                            elif suc == 3:
                                view.clear_items()
                                add_score(interaction.user.name, userRow, 5)
                                await interaction.channel.send(embed = discord.Embed(title='클러치!', 
                                description="상대방을 처치하고 스파이크를 해체했습니다.",colour=discord.Colour.blue()), 
                                view=view)
                            elif suc == 4:
                                view.clear_items()
                                add_score(interaction.user.name, userRow, 10)
                                await interaction.channel.send(embed = discord.Embed(title='팀 에이스!', 
                                description="팀원이 모두 한명씩 처치했습니다",colour=discord.Colour.blue()), 
                                view=view)
                            elif suc == 5:
                                view.clear_items()
                                add_score(interaction.user.name, userRow, 30)
                                await interaction.channel.send(embed = discord.Embed(title='에이스!', 
                                description="5명의 모든 적을 처치했습니다!! 당신은 이길 자격이 있군요!",colour=discord.Colour.blue()), 
                                view=view)
                            elif suc == 6:
                                view.clear_items()
                                add_score(interaction.user.name, userRow, 100)
                                await interaction.channel.send(embed = discord.Embed(title='퍼펙트 에이스!', 
                                description="어마어마하네요. ㅎ",colour=discord.Colour.blue()), 
                                view=view)
                    else:
                        await interaction.channel.send(f'{user}님!! 남의 오퍼레이터를 훔쳐가면 안되요!')

            else:
                await interaction.response.send_message("오늘 횟수를 다 사용하셨습니다. 내일 다시 시도해 주세요.", ephemeral = True)
        else:
            await interaction.response.send_message("회원가입 후 이용해주세요.", ephemeral = True)

    button1.callback = button_callback1
    view = View()
    view.add_item(button1)
    guild = client.get_guild(1053299673206640711)
    gi = guild.icon
    iun = interaction.user.name
    iua = interaction.user.avatar
    embed=discord.Embed(title=f'오퍼레이터 사격 게임',description="오퍼레이터를 사격해 적을 처치합니다.", color=0xf1c40f)
    embed.add_field(name=f'남은 기회', value=f'{ws.cell(userRow, c_day_count).value}번\n🚫주의 : 시작하기 버튼을 누르면 남은 기회가 소모됩니다.')
    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
    embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
    await interaction.response.send_message(embed = embed, view=view)

##################################낚시게임(beta 이후 업데이트 예정)#################################


##################################도움(Help) 명령어################################

@tree.command(name="도움말", description="봇을 사용하기 위한 도움말을 보여줍니다.")
async def help(interaction: discord.Interaction):
    guild = client.get_guild(1053299673206640711)
    gi = guild.icon
    iun = interaction.user.name
    iua = interaction.user.avatar
    embed=discord.Embed(title=f'<a:_heart:1061559734081171476> 안녕하세요? <a:_heart:1061559734081171476>',description="저는 발로란트 학교를 위한 다목적 봇이에요!", color=0xf1c40f)
    embed.add_field(name="/도움말 <a:star_1:1061559907595329568>",value=f"`봇을 사용하기 위한 도움말을 보여줘요.`",inline=False)
    embed.add_field(name="/회원가입 <a:star_2:1061559966835691590>",value=f"`오퍼레이터 사격 게임을 하기 위한 회원가입 명령어에요.`",inline=False)
    embed.add_field(name="/오퍼레이터사격 <a:star_3:1061559984955080714>",value=f"`오퍼레이터 사격 게임을 플레이하는 명령어에요.`",inline=False)
    embed.add_field(name="/내정보 <a:star_1:1061559907595329568>",value=f"`닉네임, 점수 등 본인의 정보를 확인할 수 있는 명령어에요.\n현재 개발중이에요!`:wrench:",inline=False)
    embed.add_field(name="/포인트도움말 <a:star_2:1061559966835691590>",value=f"`포인트와 관련된 도움말을 보여줍니다.`",inline=False)
    embed.add_field(name="/랭킹 <a:star_3:1061559984955080714>",value=f"`본인의 랭킹과 전체 랭킹을 보여줍니다.`",inline=False)
    embed.add_field(name="🚫주의 사항",value=f"`회원가입 후 디스코드 닉네임 변경시 원활하게 사용이 불가합니다.`",inline=False)
    embed.set_thumbnail(url=f"{gi}")
    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
    embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
    await interaction.response.send_message(embed=embed)

@tree.command(name="포인트도움말", description="포인트와 관련된 도움말을 보여줍니다.")
async def point_help(interaction: discord.Interaction):
    guild = client.get_guild(1053299673206640711)
    gi = guild.icon
    iun = interaction.user.name
    iua = interaction.user.avatar
    embed=discord.Embed(title=f'<a:_heart:1061559734081171476> 안녕하세요? <a:_heart:1061559734081171476>',description="저는 발로란트 학교를 위한 다목적 봇이에요!", color=0xf1c40f)
    embed.add_field(name="/rsp상점 <a:star_1:1061559907595329568>",value=f"<:z_rp:1070890924462317580>`Rsp를 사용할 수 있는 상점을 보여줍니다.`",inline=False)
    embed.add_field(name="/vsp상점 <a:star_2:1061559966835691590>",value=f"<:z_vp:1070890926928580618>`Vsp를 사용할 수 있는 상점을 보여줍니다.`",inline=False)
    embed.add_field(name="/rsp사용 <a:star_3:1061559984955080714>",value=f"<:z_rp:1070890924462317580>`Rsp를 사용해 상품을 구매합니다.`",inline=False)
    embed.add_field(name="/vsp사용 <a:star_1:1061559907595329568>",value=f"<:z_vp:1070890926928580618>`Vsp를 사용해 상품을 구매합니다.`",inline=False)
    embed.add_field(name="/포인트획득방법 <a:star_2:1061559966835691590>",value=f"`포인트를 획득 할 수있는 방법을 알려줍니다.`",inline=False)
    embed.set_thumbnail(url=f"{gi}")
    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
    embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
    await interaction.response.send_message(embed=embed)

@tree.command(name="포인트획득방법", description="포인트 획득 방법을 알려줍니다.")
async def point_get_help(interaction: discord.Interaction):
    guild = client.get_guild(1053299673206640711)
    gi = guild.icon
    iun = interaction.user.name
    iua = interaction.user.avatar
    embed=discord.Embed(title=f'💸 포인트 획득 방법 💸',description="Rsp와 Vsp의 획득 방법이에요!", color=0xf1c40f)
    embed.add_field(name="<:z_rp:1070890924462317580> Rsp",value=f"`여러가지 이벤트를 참여하면 획득 할 수 있습니다.\n내전참여시 100Rsp, 일주일 출석 완료시 100Rsp,\nMee6봇 레벨 10레벨부터 5레벨마다 1000rsp\n출석은 오퍼레이터 사격게임에 참가하면 인정됨.`",inline=False)
    embed.add_field(name="<:z_vp:1070890926928580618> Vsp",value=f"`서버 후원시 10원당 1Vsp,\n서버 부스트 사용시 개당 250Vsp\n\n그 외 이벤트로 Rsp나 Vsp를 지급할 수 있습니다.`",inline=False)
    embed.set_thumbnail(url=f"{gi}")
    embed.set_author(name=f"{iun}", icon_url=f"{iua}")
    embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
    await interaction.response.send_message(embed=embed)

#####################################시간 체크 테스트##################################

def day_count_reset(all_user):
    loadFile()
    for row1 in range(2, 2 + int(all_user)):
        ws.cell(row1, c_day_count).value = 1
        print(row1, "번의 day_count 초기화 완료")
    print("초기화가 끝났습니다.")
    print("------------------------------\n")
    saveFile()

def check_daytime():
    now = datetime.datetime.now()
    loadFile()
    userNum1 = checkUserNum()
    print("등록된 유저수: ", userNum1)
    print(" ")
    print ("날짜 : ", now.strftime("%Y / %m / %d"))
    print(" ")
    print("하루가 지났습니다.")
    print("day_count를 초기화합니다.")
    print("------------------------------\n")
    day_count_reset(int(userNum1))
    saveFile()

@sched.scheduled_job('cron', hour='00', minute='00', id='day')
def hour_check():
    print(f'Now Time : {time.strftime("%H:%M:%S")}, Initiallizing Game_play_count...')
    check_daytime()

sched.start()

###################################### 내정보 확인 하기 ####################################

@tree.command(name="내정보", description="닉네임, 점수 등 본인의 정보를 확인할 수 있는 명령어에요.")
async def my_Information(interaction: discord.Interaction):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        guild = client.get_guild(1053299673206640711)
        gi = guild.icon
        iun = interaction.user.name
        iua = interaction.user.avatar
        userRowRank = getRank(userRow)
        embed=discord.Embed(title=f'<a:check_8:1060790808388841522> {interaction.user.name}님의 프로필 <a:check_8:1060790808388841522>',description="ㅤ", color= 0xf1c40f)
        embed.add_field(name="<:star1:1061560017746137119> 닉네임",value=f"{interaction.user.name}",inline=False)
        embed.add_field(name="<:star2:1061560043914399814> 점수",value=f"{ws.cell(userRow, c_score).value}점",inline=False)
        embed.add_field(name="<:star3:1061560074335694888> 남은 기회",value=f"{ws.cell(userRow, c_day_count).value}번",inline=False)
        embed.add_field(name="<:star4:1061560100311023697> 랭킹",value=f"{userRowRank}위",inline=False)
        embed.add_field(name="<:star5:1061560147085905930> Rsp",value=f"<:z_rp:1070890924462317580>{ws.cell(userRow, c_rsp).value}Rsp",inline=False)
        embed.add_field(name="<:star6:1061560169965817916> Vsp",value=f"<:z_vp:1070890926928580618>{ws.cell(userRow, c_vsp).value}Vsp",inline=False)
        embed.set_thumbnail(url=f"{gi}")
        embed.set_author(name=f"{iun}", icon_url=f"{iua}")
        embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
        await interaction.response.send_message(embed=embed)
    else:
        await interaction.response.send_message(f"회원가입 후 이용해주세요.", ephemeral = True) 

######################################포인트 지급#####################################

@tree.command(name="포인트지급", description="유저에게 포인트를 지급합니다.")
@app_commands.describe(username = "포인트를 지급할 유저의 이름을 입력하세요.", point = "지급할 포인트의 양을 입력하세요.")
@app_commands.choices(choices=[
    app_commands.Choice(name="Rsp", value="rsp"),
    app_commands.Choice(name="Vsp", value="vsp"),
    ])
async def give_point(interaction: discord.Interaction, username : str, point : int, choices: app_commands.Choice[str]):
    if interaction.guild:
        if interaction.user.guild_permissions.administrator:
            uc = client.get_channel(1070492357528649850)
            userExistance, userRow = checkUser_name(username)
            if userExistance:
                loadFile()
                if (choices.value == 'rsp'):
                    await interaction.response.send_message(f'{username}에게 <:z_rp:1070890924462317580>{point}Rsp를 지급하였습니다.')
                    ws.cell(userRow, c_rsp).value += point
                    print(f'{username}에게 {point}Rsp를 지급했습니다.')
                    print("------------------------------\n")
                    saveFile()
                    embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 {username}님에게 <:z_rp:1070890924462317580>{point}Rsp를 지급하였습니다.", timestamp=datetime.datetime.now(),color=0x00ff13)
                    await uc.send(embed=embed)
                if (choices.value == 'vsp'):
                    await interaction.response.send_message(f'{username}에게 <:z_vp:1070890926928580618>{point}Vsp를 지급하였습니다.')
                    ws.cell(userRow, c_vsp).value += point
                    print(f'{username}에게 {point}Vsp를 지급했습니다.')
                    print("------------------------------\n")
                    saveFile()
                    embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 {username}님에게 <:z_vp:1070890926928580618>{point}Vsp를 지급하였습니다.", timestamp=datetime.datetime.now(),color=0x00c7ff)
                    await uc.send(embed=embed)
            else:
                print(f'{username}는 존재하지 않는 유저임.')
                print("------------------------------\n")
                await interaction.response.send_message(f'{username}는 존재하지 않는 유저입니다.')
        else:
            await interaction.response.send_message(f'포인트 제도를 악용하려 하지 마세요.')
    else:
        await interaction.response.send_message(f'심각한 오류가 발생하였습니다.')

#######################################포인트 상점##########################################

@tree.command(name="rsp상점", description="Rsp를 사용할 수 있는 상점을 엽니다.")
async def point_shop_rsp(interaction: discord.Interaction):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        await interaction.response.send_message(f'ㅤ', ephemeral = True)
        loadFile()
        guild = client.get_guild(1053299673206640711)
        gi = guild.icon
        embed=discord.Embed(title=f'<:z_rp:1070890924462317580> Rsp상점 <:z_rp:1070890924462317580>',description="Rsp를 사용 할 수 있는 상점입니다.\nRsp 획득 방법은 포인트도움말을 참고하세요!", color=0x1abc9c)
        embed.add_field(name="🚫경고 차감권",value=f"1000Rsp<:z_rp:1070890924462317580>",inline=True)
        embed.add_field(name="<:game_valorant:1060783071772823662> 닉네임 장식",value=f"3000Rsp<:z_rp:1070890924462317580>",inline=True)
        embed.add_field(name="🌸 Episode 6 Act 1 Master Emblem",value=f"10000Rsp<:z_rp:1070890924462317580>",inline=True)
        embed.set_author(name=f"Valorant School", icon_url=f"{gi}")
        embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
        msg = await interaction.channel.send(embed=embed)
        await msg.add_reaction("❌")
        @client.event
        async def on_reaction_add(reaction, user):
            if user.bot == 1:
                return None
            if str(reaction.emoji) == "❌":
                await msg.delete()
            saveFile() 
    else:
        await interaction.response.send_message(f'회원가입 후 이용해주세요.', ephemeral = True)

@tree.command(name="vsp상점", description="Vsp를 사용할 수 있는 상점을 엽니다.")
async def point_shop_vsp(interaction: discord.Interaction):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        await interaction.response.send_message(f'ㅤ', ephemeral = True)
        loadFile()
        guild = client.get_guild(1053299673206640711)
        gi = guild.icon
        embed=discord.Embed(title=f'<:z_vp:1070890926928580618> Vsp 상점 <:z_vp:1070890926928580618>',description="Vsp를 사용 할 수 있는 상점입니다.\nVsp 획득 방법은 포인트도움말을 참고하세요!", color=0x1abc9c)
        embed.add_field(name="미정 상품",value=f"1000Vsp<:z_vp:1070890926928580618>",inline=True)
        embed.add_field(name="커스텀 닉네임 장식",value=f"1000Vsp<:z_vp:1070890926928580618>",inline=True)
        embed.add_field(name="이벤트 상품",value=f"500Vsp<:z_vp:1070890926928580618>",inline=True)
        embed.set_author(name=f"Valorant School", icon_url=f"{gi}")
        embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
        msg = await interaction.channel.send(embed=embed)
        await msg.add_reaction("❌")
        @client.event
        async def on_reaction_add(reaction, user):
            if user.bot == 1:
                return None
            if str(reaction.emoji) == "❌":
                await msg.delete()
            saveFile() 
    else:
        await interaction.response.send_message(f'회원가입 후 이용해주세요.', ephemeral = True)

@tree.command(name="rsp사용", description="Rsp를 사용하여 원하는 상품을 구매합니다.")
@app_commands.choices(choices=[
    app_commands.Choice(name="경고 차감권", value="경고차감권"),
    app_commands.Choice(name="닉네임 장식", value="닉네임장식"),
    app_commands.Choice(name="Episode 6 Act 1 Master Emblem", value="Episode 6 Act 1 Master Emblem"),
    ])
async def buy_item_rsp(interaction: discord.Interaction, choices: app_commands.Choice[str]):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        uc = client.get_channel(1070492357528649850)
        loadFile()
        if (choices.value == '경고차감권'):
            if ws.cell(userRow, c_rsp).value >= 1000:
                await interaction.response.send_message(f'{interaction.user.name}님이 경고 차감권을 구매하셨습니다.')
                ws.cell(userRow, c_rsp).value -= 1000
                embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 경고차감권을 구매하셨습니다.\n{interaction.user.name}님의 보유 Rsp : {ws.cell(userRow, c_rsp).value}", timestamp=datetime.datetime.now(),color=0x00ff13)
                await uc.send(embed=embed)
                saveFile()
            else:
                await interaction.response.send_message(f'{interaction.user.name}님의 Rsp가 부족합니다.')
        elif (choices.value == '닉네임장식'):
            if ws.cell(userRow, c_rsp).value >= 3000:
                await interaction.response.send_message(f'{interaction.user.name}님이 닉네임 장식을 구매하셨습니다')
                ws.cell(userRow, c_rsp).value -= 3000
                embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 닉네임 장식을 구매하셨습니다.\n{interaction.user.name}님의 보유 Rsp : {ws.cell(userRow, c_rsp).value}", timestamp=datetime.datetime.now(),color=0x00ff13)
                await uc.send(embed=embed)
                saveFile()
            else:
                await interaction.response.send_message(f'{interaction.user.name}님의 Rsp가 부족합니다.')
        elif (choices.value == 'Episode 6 Act 1 Master Emblem'):
            if ws.cell(userRow, c_rsp).value >= 10000:
                await interaction.response.send_message(f'{interaction.user.name}님이 Episode 6 Act 1 Master Emblem을 구매하셨습니다.')
                ws.cell(userRow, c_rsp).value -= 10000
                embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 Episode 6 Act 1 Master Emblem을 구매하셨습니다.\n{interaction.user.name}님의 보유 Rsp : {ws.cell(userRow, c_rsp).value}", timestamp=datetime.datetime.now(),color=0x00ff13)
                await uc.send(embed=embed)
                saveFile()
            else:
                await interaction.response.send_message(f'{interaction.user.name}님의 Rsp가 부족합니다.')                
    else:
        await interaction.response.send_message(f'회원가입 후 이용해주세요.', ephemeral = True)

@tree.command(name="vsp사용", description="Vsp를 사용하여 원하는 상품을 구매합니다.")
@app_commands.choices(choices=[
    app_commands.Choice(name="미정 상품", value="미정 상품"),
    app_commands.Choice(name="커스텀 닉네임 장식", value="커스텀 닉네임 장식"),
    app_commands.Choice(name="이벤트 상품", value="이벤트 상품"),
    ])
async def buy_item_vsp(interaction: discord.Interaction, choices: app_commands.Choice[str]):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        uc = client.get_channel(1070492357528649850)
        loadFile()
        if (choices.value == '미정 상품'):
            if ws.cell(userRow, c_vsp).value >= 1000:
                await interaction.response.send_message(f'{interaction.user.name}님이 미정 상품을 구매하셨습니다.')
                ws.cell(userRow, c_vsp).value -= 1000
                embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 미정 상품을 구매하셨습니다.\n{interaction.user.name}님의 보유 Vsp : {ws.cell(userRow, c_vsp).value}", timestamp=datetime.datetime.now(),color=0x00c7ff)
                await uc.send(embed=embed)
                saveFile()
            else:
                await interaction.response.send_message(f'{interaction.user.name}님의 Vsp가 부족합니다.')
        elif (choices.value == '커스텀 닉네임 장식'):
            if ws.cell(userRow, c_vsp).value >= 1000:
                await interaction.response.send_message(f'{interaction.user.name}님이 커스텀 닉네임 장식을 구매하셨습니다\n티어 이모지 제외 원하시는 사진이나 이모지를 부시#0150 DM으로 전달해주세요!')
                ws.cell(userRow, c_vsp).value -= 1000
                embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 커스텀 닉네임 장식을 구매하셨습니다.\n{interaction.user.name}님의 보유 Vsp : {ws.cell(userRow, c_vsp).value}", timestamp=datetime.datetime.now(),color=0x00c7ff)
                await uc.send(embed=embed)
                saveFile()
            else:
                await interaction.response.send_message(f'{interaction.user.name}님의 Vsp가 부족합니다.')
        elif (choices.value == '이벤트 상품'):
            if ws.cell(userRow, c_vsp).value >= 500:
                await interaction.response.send_message(f'{interaction.user.name}님이 이벤트 상품을 구매하셨습니다.')
                ws.cell(userRow, c_vsp).value -= 500
                embed=discord.Embed(title="Debug Info",description=f"{interaction.user.name}님이 이벤트 상품을 구매하셨습니다.\n{interaction.user.name}님의 보유 Vsp : {ws.cell(userRow, c_vsp).value}", timestamp=datetime.datetime.now(),color=0x00c7ff)
                await uc.send(embed=embed)
                saveFile()
            else:
                await interaction.response.send_message(f'{interaction.user.name}님의 Vsp가 부족합니다.')
    else:
        await interaction.response.send_message(f'회원가입 후 이용해주세요.', ephemeral = True)


#####################################랭킹#######################################

def ranking():
    print("ranking")

    loadFile()

    userRanking =  {}
    userNum = checkUserNum()

    print("등록된 유저수: ", userNum)
    print("")

    print("랭킹 집계중")

    for row in range(2, 2+userNum):
        _name = ws.cell(row, c_name).value
        _score = ws.cell(row, c_score).value
        userRanking[_name] = _score

    print("랭킹 집계 완료")
    a = sorted(userRanking.items(), reverse=True, key=lambda item:item[1])
    result = []
    for items in a:
        result.append(items[0])
        result.append(items[1])
    print(result)
    print("")
    print("------------------------------\n")

    return result

def getRank(_row):
    print("getRank")
    user = ws.cell(_row, c_name).value
    print(user, "의 랭킹 조사")
    rank = ranking()

    result = int(rank.index(user)/2)+1
    print(user, "의 랭킹: ",result, "위")

    return result

@tree.command(name="랭킹", description="개인 랭킹과 전체 랭킹을 보여줍니다.")
async def show_ranking(interaction: discord.Interaction):
    userExistance, userRow = checkUser(interaction.user.name, interaction.user.id)
    if userExistance:
        rank = ranking()
        userRowRank = getRank(userRow)
        guild = client.get_guild(1053299673206640711)
        gi = guild.icon
        iun = interaction.user.name
        iua = interaction.user.avatar
        embed = discord.Embed(title = "🏆 랭킹 등수 🏆", description = "오퍼레이터 사격 게임의 랭킹 등수입니다.", color = 0x1abc9c)
        embed.add_field(name =f"<a:check_8:1060790808388841522> {interaction.user.name}님의 랭킹", value =f"점수 : {ws.cell(userRow, c_score).value}\n 등수 : {userRowRank}위", inline=False)
        embed.set_author(name=f"{iun}", icon_url=f"{iua}")
        embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
        await interaction.response.send_message(embed=embed, ephemeral = True)
        embed = discord.Embed(title = "🏆 점수 랭킹 🏆", description = "오퍼레이터 사격 게임의 점수 랭킹입니다.", color = 0x1abc9c)
        for i in range(0,5):
            if i%2 == 0:
                name = rank[i]
                score = rank[i+1]
                if int(i/2+1) == 1:
                    embed.add_field(name = "<:9Raddiant:1054692592459198465> 1위 "+name, value ="점수 : "+str(score), inline=False)
                if int(i/2+1) == 2:
                    embed.add_field(name = "<:8Immortal:1054692575363223562> 2위 "+name, value ="점수 : "+str(score), inline=False)
                if int(i/2+1) == 3:
                    embed.add_field(name = "<:7Ascendant:1054692555964567613> 3위 "+name, value ="점수 : "+str(score), inline=False)
        embed.set_footer(text=f"버그 발생 시, Yena#0001 또는 부시#0150로 문의해주세요.\nMade with ♥ By Yena#0001(492899510255747072)\nMade with ♥ By 부시#0150(536026459471609906)")
        embed.set_author(name=f"Valorant School", icon_url=f"{gi}")
        msg = await interaction.channel.send(embed=embed)
        await msg.add_reaction("❌")
        @client.event
        async def on_reaction_add(reaction, user):
            if user.bot == 1:
                return None
            if str(reaction.emoji) == "❌":
                await msg.delete()
    else:
        await interaction.response.send_message("회원가입 후 이용해주세요.", ephemeral = True)

#🥇🥈🥉

######################################토큰#######################################

client.run(token)
